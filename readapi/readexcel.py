import requests
import openpyxl
import json
import datetime
import re
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)     #忽略https认证警告

wb = openpyxl.load_workbook('apis.xlsx')
ws = wb['Sheet1']
data = []
for i in  range(2,ws.max_row+1):
    dict = {}
    header = {
        "Essc-Referer": "http://127.0.0.1",
        "User-Agent": "Mozilla/5.0 (iPhone; CPU iPhone OS 12_2 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Mobile/15E148 JinbaoxinClient",
        "Accept-Encoding": "gzip, deflate",
        "Accept-Language": "zh-cn",
        "Accept": "application/json, text/plain, */*",
        "X-Requested-With": "XMLHttpRequest",
        "Cache-Control": "no-cache",
        "Accept-Encoding": "gzip, deflate"
    }
    for j in range(1,ws.max_column+1):
        key = ws.cell(1,j).value
        dict[key] = ws.cell(i,j).value
        if ws.cell(1,ws.max_column).value in key:       #判断第一行标题已轮训完
            if dict['needKey'] is not None:
                dict['parameter'] = dict['parameter'].replace(dict['replaceKey'], globals()[dict['needKey']])

            if dict['needHeader'] is not None:
                dict['addHeader'] = dict['addHeader'].replace(dict['replaceHeader'],globals()[dict['needHeader']])

            if dict['addHeader'] is not None:
                header.update(eval(dict['addHeader']))       #dict['addHeader']类型为字符串，通过eval转化成字典类型，再通过update方法添加到header

            if dict['function'] == 'post':
                res = requests.post(url=dict['api'], data=dict['parameter'].encode('utf-8'),
                                        headers=header,verify=False)  # verify=False 关闭https认证
                print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')+"--"+dict['apiName']+"--请求--"+dict['parameter'])
                print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')+"--"+dict['apiName']+"--响应--"+res.text)

            if dict['function'] == 'get':
                res = requests.get(url=dict['api'], params=dict['parameter'], headers=header, verify=False)  # verify=False 关闭https认证
                print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S') + "--" + dict['apiName'] + "--请求--" + dict['parameter'])
                print(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S') + "--" + dict['apiName'] + "--响应--" + res.text)

            if dict['resultKey'] is not None:
                resultkey = re.findall(dict['resultKey'] + r'":"(.*?)"', res.text)    #正则提取
                globals()[dict['resultKey']] = ''.join(resultkey[0])  # join将result由列表转化成字符串,如果匹配到多个值，只取第一个

            #添加断言
            exp = re.findall(r'"msg":"(.*?)"', res.text)
            exp = ''.join(exp[0])
            if  dict['expValue'] != exp:
                print(dict['api']+"   失败")

    data.append(dict)

print(data)
